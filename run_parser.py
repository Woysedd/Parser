import asyncio
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import requests
import os

HEADERS = ["Название устройства", "re:luxon", "re:premium", "Like Store", "Prostore", "KingStore Уфа"]

FILL_YELLOW = PatternFill(start_color="FFF2CC", fill_type="solid") 
FILL_GREEN = PatternFill(start_color="E2EFDA", fill_type="solid")  
FILL_GREY = PatternFill(start_color="F2F2F2", fill_type="solid")   
FILL_PINK = PatternFill(start_color="FCE4D6", fill_type="solid")   
FILL_HEADER = PatternFill(start_color="A9D08E", fill_type="solid")

FONT_BOLD = Font(name="Arial", size=11, bold=True)
FONT_REGULAR = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
)

PRO_COLORS = ["Black", "White", "Natural", "Gold"]
PRO_SIZES = ["256Gb", "512Gb", "1Tb"]
SIM_TYPES = ["eSIM", "SIM+eSIM"]

BASE_DEVICES = [
    ("iPhone 13 128Gb", FILL_YELLOW, "13", "128", ""),
    ("iPhone 15 128Gb", FILL_YELLOW, "15", "128", ""),
    ("iPhone 16 128Gb", FILL_GREEN, "16", "128", ""),
    ("iPhone 16 256Gb", FILL_GREEN, "16", "256", ""),
    ("iPhone Air 256Gb", FILL_GREEN, "air", "256", ""),
    ("iPhone 17 128Gb eSIM", FILL_GREY, "17", "128", "esim"),
    ("iPhone 17 256Gb eSIM", FILL_GREY, "17", "256", "esim"),
    ("iPhone 17 256Gb SIM+eSIM", FILL_GREY, "17", "256", "sim"),
    ("iPhone 17 512Gb SIM+eSIM", FILL_GREY, "17", "512", "sim"),
]

DEVICES_MATRIX = list(BASE_DEVICES)
for model in ["iPhone 17 Pro", "iPhone 17 Pro Max"]:
    m_short = "17 pro max" if "max" in model.lower() else "17 pro"
    for size in PRO_SIZES:
        s_short = size.replace("Gb", "")
        for color in PRO_COLORS:
            for sim in SIM_TYPES:
                sim_short = "sim" if "sim+" in sim.lower() else "esim"
                full_name = f"{model} {size} {color} {sim}"
                DEVICES_MATRIX.append((full_name, FILL_PINK, m_short, s_short, sim_short))

def clean_price(raw_text):
    if not raw_text: return "По запросу"
    digits = "".join([c for c in raw_text if c.isdigit()])
    if not digits or len(digits) < 4: return "По запросу"
    if len(digits) > 6: digits = digits[:6]
    return f"{digits[:-3]}.{digits[-3:]}"

async def fetch_site_data(context, url, item_sel, title_sel, price_sel):
    """Скачивает все товары с сайта за один заход в один клик"""
    data = []
    page = await context.new_page()
    try:
        await page.goto(url, timeout=45000, wait_until="domcontentloaded")
        await asyncio.sleep(3)
        items = await page.query_selector_all(item_sel)
        for item in items:
            t_el = await item.query_selector(title_sel)
            p_el = await item.query_selector(price_sel)
            if t_el and p_el:
                t_text = (await t_el.text_content()).lower()
                p_text = await p_el.text_content()
                data.append((t_text, clean_price(p_text)))
    except Exception as e:
        print(f"Ошибка скачивания {url}: {e}")
    finally:
        await page.close()
    return data

def find_price_in_cache(cache, m_num, mem, sim_type):
    """Мгновенно ищет нужную модель в скачанном кэше без запросов в сеть"""
    for title_text, price in cache:
        if m_num in title_text and mem in title_text:
            if "max" in m_num and "max" not in title_text: continue
            if "max" not in m_num and "max" in title_text: continue
            if "pro" in title_text and "pro" not in m_num: continue
            if sim_type == "sim" and "esim" in title_text and "sim+" not in title_text: continue
            if sim_type == "esim" and "esim" not in title_text: continue
            return price
    return "По запросу"

async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        
        print("🚀 Одновременно скачиваем каталоги всех конкурентов...")
        # Запускаем параллельный сбор данных со всех 5 сайтов разом
        tasks = [
            fetch_site_data(context, "https://re-luxe42.ru/iphone/iphone-new", ".product-thumb", ".caption a", ".price"),
            fetch_site_data(context, "https://repremium.ru/kemerovo/catalog/apple/iphone/", ".catalog-item", ".item-title", ".price_val"),
            fetch_site_data(context, "https://kemerovo.lstore.ru/catalog/iphone_1/", ".product-card", ".product-card__title", ".product-card__price-current"),
            fetch_site_data(context, "https://prostore-shop.ru/catalog_iphone", ".t-store__card", ".t-store__card__title", ".t-store__card__price-value"),
            fetch_site_data(context, "https://kingstore.link/catalog/iphone/", ".catalog-item", ".item-title", ".price_val")
        ]
        
        caches = await asyncio.gather(*tasks)
        c_luxon, c_premium, c_like, c_pro, c_king = caches
        
        print("📊 Формируем итоговую Excel-таблицу...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Срез цен"
        ws.views.sheetView[0].showGridLines = True
        
        ws.append(HEADERS)
        for c_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.fill = FILL_HEADER
            cell.font = FONT_BOLD
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
        
        for dev_name, fill, m_num, mem, sim_type in DEVICES_MATRIX:
            p_luxon = find_price_in_cache(c_luxon, m_num, mem, sim_type)
            p_premium = find_price_in_cache(c_premium, m_num, mem, sim_type)
            p_like = find_price_in_cache(c_like, m_num, mem, sim_type)
            p_pro = find_price_in_cache(c_pro, m_num, mem, sim_type)
            p_king_ufa = find_price_in_cache(c_king, m_num, mem, sim_type)
            
            ws.append([dev_name, p_luxon, p_premium, p_like, p_pro, p_king_ufa])
            
            r_idx = ws.max_row
            for c_idx in range(1, len(HEADERS) + 1):
                c = ws.cell(row=r_idx, column=c_idx)
                c.fill = fill
                c.font = FONT_REGULAR
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 18)
        
        date_str = datetime.datetime.now().strftime("%Y-%m-%d")
        file_name = f"Прайс_Конкурентов_{date_str}.xlsx"
        wb.save(file_name)
        await browser.close()
        
        token = os.environ['TELEGRAM_BOT_TOKEN']
        chat_ids = os.environ['TELEGRAM_CHAT_ID'].split(',')
        url_tg = f"https://api.telegram.org/bot{token}/sendDocument"
        
        for chat_id in chat_ids:
            chat_id = chat_id.strip()
            if chat_id:
                with open(file_name, "rb") as f:
                    requests.post(url_tg, data={"chat_id": chat_id, "caption": f"📊 Мониторинг рынка iPhone на {date_str} готов!"}, files={"document": f})

if __name__ == "__main__":
    asyncio.run(main())
